from functools import wraps
import json, io
from datetime import datetime
import os
from flask import (
    Blueprint,
    abort,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    jsonify,
    url_for,
)
from flask.views import MethodView
from flask_login import login_required, current_user

from openpyxl import load_workbook
from fpdf import FPDF
import magic

from .models import Admin, Employee, Team
from .utils import (
    get_admins,
    create_employee,
    camel_case,
    hexid,
    validate_employee,
    create_excel_from_data,
    datetime_stamp,
)
from . import db

views = Blueprint("views", __name__)


def notsuspended_required(func):
    @wraps(func)
    @login_required
    def wrapper(*args, **kwargs):
        if current_user.is_suspended:
            return render_template("auth/admin_susp.html")
        return func(*args, **kwargs)

    return wrapper


class Verify(MethodView):
    def get(self):
        return render_template("views/verify.html")

    def post(self):
        data = json.loads(request.data)
        employee_alpha_code = data.get("employeeAlphaCode", None)
        team_alpha_code = data.get("teamAlphaCode", None)

        if employee_alpha_code:
            employee = Employee.query.filter_by(alpha_code=employee_alpha_code).first()
            if employee:
                return jsonify(employee.to_dict)
            return jsonify({}), 404
        elif team_alpha_code:
            team = Team.query.filter_by(alpha_code=team_alpha_code).first()
            if team:
                return jsonify(team.to_dict)
            return jsonify({}), 404

        return abort(400)


views.add_url_rule("/", view_func=Verify.as_view("verify"))


class AdminHome(MethodView):
    decorators = [login_required, notsuspended_required]

    def get(self):
        admins = []
        if current_user.is_super:
            admins = get_admins(request, current_user)

        employees = Employee.query.all()
        return render_template(
            "views/admin_home.html", admins=admins, employees=employees
        )

    def post(self):
        data = json.loads(request.data)
        res = create_employee(data)
        if res["error"]:
            return jsonify({"message": res["message"], "cate": "danger"})
        return jsonify({"message": "Employee Saved Successfully", "cate": "success"})

    def put(self):
        data = json.loads(request.data)
        err = validate_employee(data)
        if err != "":
            return jsonify({"message": err, "cate": "danger"})

        employee = Employee.query.get(int(data["id"]))
        if not employee:
            return abort(404)

        employee.alpha_code = data["alphaCode"]
        employee.first_name = data["firstName"]
        employee.last_name = data["lastName"]
        employee.gender = data["gender"]
        employee.job_title = data["jobTitle"]
        employee.department = data["department"]
        employee.date_of_hire = datetime.strptime(data["dateOfHire"], "%Y-%m-%d")

        db.session.commit()
        return jsonify({**data, "cate": "success"})

    def delete(self):
        data = json.loads(request.data)
        leader_ids = []

        for team in Team.query.all():
            leader_ids.append(team.leader_id)

        found_leaders = []
        for emp_id in data["employees"]:
            emp_id = int(emp_id)
            if emp_id not in leader_ids:
                employee = Employee.query.get(emp_id)
                db.session.delete(employee)
            else:
                found_leaders.append(emp_id)

        db.session.commit()
        data["found_leaders"] = found_leaders

        return jsonify(data)


views.add_url_rule("/admin/", view_func=AdminHome.as_view("admin_home"))


class ExportEmployeeExcel(MethodView):
    decorators = [login_required]

    def get(self):
        """Endpoint to download all employee data as an Excel file."""
        # Query all employees from the database
        employees = Employee.query.all()

        # Generate Excel file
        excel_file = create_excel_from_data(employees)

        filename = f"emp-{datetime_stamp()}.xlsx"
        # Send the file as a downloadable response
        return send_file(
            excel_file,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )


views.add_url_rule(
    "/export/emp/xl", view_func=ExportEmployeeExcel.as_view("export_emp_xl")
)


class ExportEmployeePdf(MethodView):
    def get(self):
        # Query all employees
        employees = Employee.query.all()

        # Create a PDF object
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # Set title
        pdf.set_font("Arial", size=12, style="B")
        pdf.cell(200, 10, txt="Employee Data", ln=True, align="C")

        # Add table headers
        pdf.set_font("Arial", size=10)
        headers = [
            "Id",
            "Alpha Code",
            "First Name",
            "Last Name",
            "Gender",
            "Job Title",
            "Department",
            "Date Of Hire",
        ]

        # Calculate column widths by finding the longest content
        col_widths = []
        for header in headers:
            max_width = pdf.get_string_width(header) + 6  # Start with header width
            for employee in employees:
                content = ""
                if header == "Id":
                    content = str(employee.id)
                elif header == "Alpha Code":
                    content = employee.alpha_code
                elif header == "First Name":
                    content = employee.first_name
                elif header == "Last Name":
                    content = employee.last_name
                elif header == "Gender":
                    content = employee.gender
                elif header == "Job Title":
                    content = employee.job_title
                elif header == "Department":
                    content = employee.department
                elif header == "Date Of Hire":
                    content = employee.date_of_hire.strftime("%Y-%m-%d")

                # Update max width based on content
                max_width = max(max_width, pdf.get_string_width(content) + 6)

            col_widths.append(max_width)

        # Print headers with auto-calculated widths
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 10, header, border=1, align="C")
        pdf.ln()

        # Add employee data rows
        for employee in employees:
            row_data = [
                str(employee.id),
                employee.alpha_code,
                employee.first_name,
                employee.last_name,
                employee.gender,
                employee.job_title,
                employee.department,
                employee.date_of_hire.strftime("%Y-%m-%d"),
            ]
            for i, data in enumerate(row_data):
                pdf.cell(col_widths[i], 10, data, border=1)
            pdf.ln()

        filename = f"emp-{datetime_stamp()}.pdf"
        tmp_folder = "app/static/tmp/"
        filepath = tmp_folder + filename

        pdf.output(filepath)

        with open(filepath, "rb") as f:
            # Save the xl file to a BytesIO stream
            buffer = io.BytesIO(f.read())
        buffer.seek(0)

        # clear the file from the static/tmp folder
        os.remove(filepath)

        return send_file(
            buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename,
        )


views.add_url_rule(
    "/export/emp/pdf", view_func=ExportEmployeePdf.as_view("export_emp_pdf")
)


class ImportExcel(MethodView):
    decorators = [login_required]

    def post(self, *args):
        # Check if a file is part of the request
        if "spreadsheetml" not in magic.from_buffer(request.data, mime=True):
            return jsonify({"message": "Invaid Excel File", "cate": "danger"})

        file_obj = io.BytesIO(request.data)
        wb = load_workbook(file_obj, data_only=True)
        ws = wb.active

        expected_cols = [
            "Alpha Code",
            "First Name",
            "Last Name",
            "Gender",
            "Job Title",
            "Department",
            "Date Of Hire",
        ]
        actual_cols = [cell.value.title() for cell in ws[1]]

        for col in expected_cols:
            if col not in actual_cols:
                return jsonify(
                    {"message": f"column ({col}) is missing", "cate": "warning"}
                )

        employees = [
            {camel_case(actual_cols[i]): cell.value for i, cell in enumerate(row)}
            for row in ws.rows
        ][1:]
        for employee in employees:
            res = create_employee(employee)
            if res["error"]:
                return jsonify({"message": res["message"], "cate": "danger"})

        return jsonify(
            {
                "message": "File uploaded successfully",
                "cate": "success",
                "data": employees,
            }
        )


views.add_url_rule("/import/excel/", view_func=ImportExcel.as_view("import_excel"))


class ManageTeams(MethodView):
    decorators = [login_required]

    def get(self):
        admins = get_admins(request, current_user)
        nonteamed_employees = Employee.query.filter_by(team_id=None).all()
        teams = Team.query.all()
        return render_template(
            "views/manage_teams.html",
            admins=admins,
            nonteamed_employees=nonteamed_employees,
            teams=teams,
        )

    def post(self):
        data = json.loads(request.data)

        team_name = data["teamName"]
        alpha_code = data["alphaCode"]
        leader_id = int(data["leader"])
        activities = data["activities"]
        member_ids = data["members"]

        if len(team_name) and leader_id:
            team = Team.query.filter_by(name=team_name).first()
            if team:
                flash(f"Team with name <{team_name}> Already Exists", cate="danger")
            else:
                alpha_code = alpha_code.upper() if alpha_code else hexid().upper()
                team = Team(
                    name=team_name,
                    leader_id=leader_id,
                    activities=activities,
                    alpha_code=alpha_code,
                )
                db.session.add(team)

                for m_id in member_ids:
                    employee = Employee.query.get(int(m_id))
                    if employee:
                        employee.team_id = team.id

                leader = Employee.query.get(int(leader_id))
                if leader:
                    leader.team_id = team.id
                    flash(f"Team <{team.name}> Created Successfully", "success")
                db.session.commit()
                team.alpha_code += str(team.id)
                db.session.commit()

        return jsonify({})

    def put(self):
        data = json.loads(request.data)
        new_team_name = data["newTeamName"]
        team = Team.query.get(int(data["teamId"]))

        if team and len(new_team_name) >= 6:
            team_name = team.name
            team.name = new_team_name
            db.session.commit()

            return {
                "message": f"Team Name Updated Successfully From <{team_name}> To <{new_team_name}>",
                "cate": "success",
            }
        return {"message": "Team Name Failed To Update", "cate": "danger"}

    def delete(self):
        data = json.loads(request.data)
        team = Team.query.get(int(data["teamId"]))
        if team:
            members = team.members
            for member in members:
                member.team_id = None
            db.session.delete(team)
            db.session.commit()
            return jsonify({"teamId": data["teamId"]})

        return jsonify({})


views.add_url_rule("/manage/teams/", view_func=ManageTeams.as_view("manage_teams"))


class ManageAdmins(MethodView):
    decorators = [login_required]

    def get(self):
        if not current_user.is_super:
            return abort(404)
        admins = get_admins(request, current_user)
        return render_template("views/manage_admins.html", admins=admins)

    def delete(self):
        data = json.loads(request.data)
        admin = Admin.query.get(data.get("adminId", 0))
        if admin:
            db.session.delete(admin)
            db.session.commit()
        return jsonify({})

    def put(self):
        data = json.loads(request.data)
        action = data.get("action", "")
        admin_id = data.get("adminId", 0)

        admin = Admin.query.get(admin_id)
        if not admin or action not in ["suspend", "resume"]:
            return abort(404)

        if action == "suspend":
            admin.is_suspended = True
            flash(f"Admin < {admin.fullname} > Successfully Suspened", "success")
        else:
            admin.is_suspended = False
            flash(f"Admin < {admin.fullname} > Successfully Resumed", "success")

        db.session.commit()
        return jsonify({})


views.add_url_rule("/manage/admins/", view_func=ManageAdmins.as_view("manage_admins"))
